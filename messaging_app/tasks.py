from celery import shared_task
from openpyxl import load_workbook
import os
from django.http import HttpResponse, FileResponse
from .models import Instance
import logging
import requests
import time
from whatsapp.settings import BASE_URL
from django.contrib import messages
from django.utils import timezone
from messaging_app.models import Log
from datetime import datetime
#  python -m celery -A whatsapp  worker -l info -P eventlet

logging.basicConfig(filename='celery.log', level=logging.INFO)

@shared_task()
def send_message(excel_path,instance_id,log_id):
    start_time = time.time()
    print(start_time)
    logging.info(instance_id)
    instance=Instance.objects.get(id=instance_id)
    
    send_message_url= BASE_URL+f"/message/text?key={instance.instance_key}"
    owner_data = {
                    "id": instance.phone_number,
                    "message": "Sending Message initiated"
                }
    response=requests.post(send_message_url, data=owner_data , verify=True, timeout=None, allow_redirects=True, stream=False, proxies=None, headers=None, cookies=None, files=None, auth=None, hooks=None, json=None, params=None)

    
    if excel_path is not None:
            wb = load_workbook(excel_path)
            sheet = wb.active
            data = []
            is_first_row = True
            for row in sheet.iter_rows(values_only=True):
                if is_first_row:
                    is_first_row = False
                    continue 
                if len(row) >= 2: 
                    data.append((row[0], row[1]))
            logging.info(f"Data to be sent: {data}")
            log_update=Log.objects.get(id=log_id)
            for i in data:
                logging.info(f"Sending message for ID: {i[0]}, Message: {i[1]}")
                body_data = {
                    "id": i[0],
                    "message": i[1]
                }
                try:
                    response=requests.post(send_message_url, data=body_data , verify=True, timeout=None, allow_redirects=True, stream=False, proxies=None, headers=None, cookies=None, files=None, auth=None, hooks=None, json=None, params=None)
                    time.sleep(2)
                    logging.info(f"Request URL: {send_message_url}")
                    logging.info(f"Response text: {response.text}")
                    logging.info(f"Response status code: {response.status_code}")
                    if response.status_code == 200 or response.status_code == 201:
                        log_update.successfull=True
                    
                except Exception as e:
                    logging.error(f"Exception occurred: {e}")
                    log_update.successfull=False

            logging.info(datetime.now())
            log_update.ended_at = datetime.now()
            end_time = time.time()
            
            logging.info(log_update.ended_at)
            duration = end_time - start_time
            log_update.duration=duration
            print(duration)
            
            log_update.ended_at=timezone.now()
            
            print(log_update.duration)
            log_update.save()
            
            owner_data = {
                    "id": instance.phone_number,
                    "message": "Message sent successfully"
                }
            response=requests.post(send_message_url, data=owner_data , verify=True, timeout=None, allow_redirects=True, stream=False, proxies=None, headers=None, cookies=None, files=None, auth=None, hooks=None, json=None, params=None)

            logging.info(log_update.ended_at)


@shared_task()
def delete_all_instance():
    instances =Instance.objects.all()
    for instance in instances:
        try:
            print(instance)
            delete_url= BASE_URL+f"/instance/delete?key={instance.instance_key}"
            print(delete_url)
            instance.delete()
            delete_response = requests.delete(delete_url, verify=True, timeout=None, allow_redirects=True, stream=False, proxies=None, headers=None, cookies=None, files=None, data=None, auth=None, hooks=None, json=None, params=None)
            time.sleep(2)
        except Exception as e:
            print(e)
            pass
     